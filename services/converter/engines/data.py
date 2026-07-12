import csv
import io
import json
from xml.etree import ElementTree as ET

from .base import ConversionRoute, ConvertContext

try:
    import openpyxl
    _O = True
except Exception:
    _O = False

try:
    from fpdf import FPDF
    _FPDF = True
except Exception:
    _FPDF = False

NAME = "data"


def available() -> bool:
    return _O


def routes() -> list[ConversionRoute]:
    exts = ["csv", "json", "xlsx", "xml"]
    out: list[ConversionRoute] = []
    for s in exts:
        for t in exts:
            if s == t:
                continue
            out.append(ConversionRoute("data", s, t, NAME, _convert, f"{s.upper()} -> {t.upper()}"))
    # csv/json/xlsx/xml -> PDF via fpdf2 (tableau)
    if _FPDF:
        for s in exts:
            out.append(ConversionRoute("data", s, "pdf", NAME, _convert, f"{s.upper()} -> PDF"))
    return out


def _read_csv(path) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return list(csv.DictReader(f))


def _write_csv(rows, path) -> None:
    keys = list(_ordered_keys(rows))
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in keys})


def _read_json(path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    if isinstance(data, list):
        return [r if isinstance(r, dict) else {"value": r} for r in data]
    if isinstance(data, dict) and "rows" in data and isinstance(data["rows"], list):
        return data["rows"]
    if isinstance(data, dict):
        return [data]
    return [{"value": data}]


def _write_json(rows, path) -> None:
    path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def _read_xlsx(path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    out: list[dict] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({h: row[i] if i < len(row) else "" for i, h in enumerate(headers)})
    return out


def _write_xlsx(rows, path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    keys = list(_ordered_keys(rows))
    ws.append(keys)
    for r in rows:
        ws.append([r.get(k, "") for k in keys])
    wb.save(str(path))


def _read_xml(path) -> list[dict]:
    tree = ET.parse(path)
    root = tree.getroot()
    container = root if root.tag.endswith("data") or len(list(root)) > 1 else list(root)[0] if len(list(root)) == 1 else root
    rows: list[dict] = []
    for row_el in container.findall("row"):
        rows.append({f.get("name"): f.text for f in row_el.findall("field")})
    if not rows:
        for row_el in root.findall(".//row"):
            rows.append({f.get("name"): f.text for f in row_el.findall("field")})
    return rows


def _write_xml(rows, path) -> None:
    root = ET.Element("data")
    for r in rows:
        row_el = ET.SubElement(root, "row")
        for k, v in r.items():
            field = ET.SubElement(row_el, "field", {"name": str(k)})
            field.text = "" if v is None else str(v)
    ET.ElementTree(root).write(str(path), encoding="utf-8", xml_declaration=True)


def _write_pdf(rows, path) -> None:
    """Génère un PDF tabulaire à partir de lignes dict via fpdf2."""
    if not _FPDF:
        raise RuntimeError("fpdf2 requis pour CSV/JSON/XLSX/XML -> PDF")
    keys = list(_ordered_keys(rows))
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_font("Helvetica", size=9)

    # En-tête
    col_w = min(40, (pdf.w - 20) / max(len(keys), 1))
    pdf.set_font("Helvetica", "B", size=9)
    pdf.set_fill_color(220, 220, 220)
    for k in keys:
        label = str(k)[:20]
        pdf.cell(col_w, 7, label, border=1, fill=True)
    pdf.ln()

    # Lignes
    pdf.set_font("Helvetica", size=8)
    for r in rows:
        for k in keys:
            val = str(r.get(k, ""))[:30]
            pdf.cell(col_w, 6, val, border=1)
        pdf.ln()
    pdf.output(str(path))


def _ordered_keys(rows) -> list:
    keys: list = []
    for r in rows:
        for k in r.keys():
            if k not in keys:
                keys.append(k)
    if not keys:
        keys = ["value"]
    return keys


_READERS = {"csv": _read_csv, "json": _read_json, "xlsx": _read_xlsx, "xml": _read_xml}
_WRITERS = {"csv": _write_csv, "json": _write_json, "xlsx": _write_xlsx, "xml": _write_xml, "pdf": _write_pdf}


def _convert(ctx: ConvertContext) -> None:
    ctx.set_progress(10, "Lecture des donnees")
    rows = _READERS[ctx.source](ctx.input_path)
    if not isinstance(rows, list):
        rows = [rows]
    ctx.set_progress(60, "Ecriture")
    _WRITERS[ctx.target](rows, ctx.output_path)
    ctx.set_progress(100, "Termine")
